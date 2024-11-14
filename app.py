import io
from flask import Flask, request, jsonify, render_template
import cv2
import numpy as np
import os
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
from fractions import Fraction
import base64
from sklearn.cluster import KMeans, MiniBatchKMeans
from werkzeug.utils import secure_filename
from sklearn.utils import shuffle
import pandas as pd
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import PatternFill

app = Flask(__name__)


# Define a directory to save uploaded images
UPLOAD_FOLDER = 'static/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Define a scale for conversion (pixels per meter)
PIXELS_PER_METER = 10  # Adjust based on your specific image scale

import rasterio

def get_lat_lon_from_image(image_path):
    """Extract latitude and longitude from satellite image metadata."""
    try:
        with rasterio.open(image_path) as src:
            # Get the bounding box of the image in geographic coordinates
            bounds = src.bounds
            # Calculate the center point of the bounding box
            lat = (bounds.top + bounds.bottom) / 2
            lon = (bounds.left + bounds.right) / 2
            
            return lat, lon
    except Exception as e:
        print(f"Error extracting coordinates from image: {e}")
    
    return None, None

def convert_ifdrational(value):
    """Convert IFDRational or similar types to float."""
    if isinstance(value, (tuple, list)):
        return [convert_ifdrational(v) for v in value]
    elif isinstance(value, Fraction):
        return float(value)
    elif hasattr(value, 'numerator') and hasattr(value, 'denominator'):
        return float(value.numerator) / float(value.denominator)
    return value

def get_lat_lon_from_exif(image_path):
    """Extract latitude and longitude from image EXIF data."""
    try:
        image = Image.open(image_path)
        exif_data = image._getexif()
        if not exif_data:
            return None, None

        lat = lon = None
        for tag_id, value in exif_data.items():
            tag = TAGS.get(tag_id, tag_id)
            if tag == 'GPSInfo':
                for key in value:
                    sub_tag = GPSTAGS.get(key, key)
                    if sub_tag == 'GPSLatitude':
                        lat = value[key]
                    elif sub_tag == 'GPSLongitude':
                        lon = value[key]
                break

        if lat is not None and lon is not None:
            lat_decimal = lat[0] + lat[1] / 60 + lat[2] / 3600
            lon_decimal = lon[0] + lon[1] / 60 + lon[2] / 3600
            
            # Handle direction indicators
            if exif_data.get(1) == 'S':  # Latitude South
                lat_decimal = -lat_decimal
            if exif_data.get(3) == 'W':  # Longitude West
                lon_decimal = -lon_decimal

            return convert_ifdrational(lat_decimal), convert_ifdrational(lon_decimal)
        
    except Exception as e:
        print(f"Error extracting EXIF data: {e}")
    
    return None, None

def serialize_exif_data(exif_info):
    """Serialize EXIF data for JSON response, ensuring no bytes are included."""
    serialized_exif_data = {}
    for key, value in exif_info.items():
        if isinstance(value, (list, tuple)):
            serialized_exif_data[key] = [convert_ifdrational(v) for v in value if not isinstance(v, bytes)]
        elif isinstance(value, Fraction):
            serialized_exif_data[key] = float(value)
        elif isinstance(value, bytes):  # Skip bytes
              pass
    return serialized_exif_data

Image.MAX_IMAGE_PIXELS = None  # Disable the limit (use cautiously)


def convert_tiff_to_jpeg(tiff_path):
    jpeg_path = tiff_path.replace('.tiff', '.jpeg').replace('.tif', '.jpeg')  # Change the extension as needed
    with Image.open(tiff_path) as img:
        img.convert('RGB').save(jpeg_path, 'JPEG')
    return jpeg_path

@app.route('/')
def index():
    return render_template('index.html')

# Function to compress image
def compress_image(image_path, max_size=(1024, 1024), quality=90):
    """Compress the image to the specified max_size and quality."""
    with Image.open(image_path) as img:
        # Resize the image
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # Save to a BytesIO object
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format='JPEG', quality=quality)
        img_byte_arr.seek(0)

        # Save the compressed image back to a file
        compressed_image_path = image_path.replace('.tiff', '_compressed.jpeg').replace('.tif', '_compressed.jpeg')
        with open(compressed_image_path, 'wb') as f:
            f.write(img_byte_arr.getvalue())
    
    return compressed_image_path

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    # Sanitize filename and check file type
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    try:
        file.save(filepath)
        m = filepath

        # Check file size
       
        latitude, longitude = get_lat_lon_from_image(filepath)

        response_filename = filename
        # Check if the uploaded file is a TIFF
        if filename.endswith('.tiff') or filename.endswith('.tif'):
            # Convert TIFF to JPEG and keep the path
            jpeg_path = convert_tiff_to_jpeg(m)

            # Optionally, you can delete the original TIFF file if you no longer need it
            os.remove(m)
            m = jpeg_path  # Update filepath to the new JPEG file
            response_filename = filename.replace('.tiff', '.jpeg').replace('.tif', '.jpeg')


        # Get all EXIF data from the converted or original image
       # exif_data = get_exif_data(m)

        # Extract latitude and longitude if available

        # Prepare response
        return jsonify({
            'filename': response_filename,
            'latitude': str(latitude) if latitude is not None else None,
            'longitude': str(longitude) if longitude is not None else None,
            #**exif_data 
        }), 200
       
    except Exception as e:
        print(f"Error during file upload: {e}")
        return jsonify({'error': 'Failed to process the image.'}), 500
    
def get_exif_data(image_path):
    try:
        image = Image.open(image_path)
        exif_data = image._getexif() or {}

        exif_info = {}
        for tag_id, value in exif_data.items():
            tag = TAGS.get(tag_id, tag_id)
            exif_info[tag] = convert_ifdrational(value)  # Convert any fractions to floats
            
        # Ensure the EXIF data is serializable to JSON
        serialized_exif_data = {}
        for key, value in exif_info.items():
            # Skip GPS-related tags (GPSInfo is usually a tag key for GPS data)
            if 'GPS' in key:
                continue  # Skip any key containing 'GPS'
            
            if isinstance(value, (list, tuple)):
                serialized_exif_data[key] = [
                    convert_ifdrational(v) if not isinstance(v, bytes) else v.decode('utf-8', errors='ignore') 
                    for v in value
                ]
            elif isinstance(value, Fraction):
                serialized_exif_data[key] = float(value)
            elif isinstance(value, bytes):
                serialized_exif_data[key] = value.decode('utf-8', errors='ignore')  # Convert bytes to string
            else:
                serialized_exif_data[key] = value
        
        print(serialized_exif_data)
                
        return serialized_exif_data

    except Exception as e:
        print(f"Error extracting EXIF data: {e}")
    
    return {}


def rgb_to_hex(rgb):
    """Convert RGB color to HEX format."""
    return '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])





@app.route('/extract_patterns', methods=['POST'])
def extract_patterns():
    data = request.json
    filename = data.get('filename')
    roi = data.get('roi')

    if not filename or not roi:
        return jsonify({'error': 'Missing filename or ROI data.'}), 400

    image_path = os.path.join(UPLOAD_FOLDER, filename)
    image = cv2.imread(image_path)
    
    with rasterio.open(image_path) as src:
        i = src.read()  # Read the image data
        transform = src.transform  # Get the affine transform for coordinate conversion
        
        # Extract coordinates of the ROI corners
        x_min, y_min = transform * (roi['x'], roi['y'])
        x_max, y_max = transform * (roi['x'] + roi['width'], roi['y'] + roi['height'])

        # Create a list of corner coordinates (longitude, latitude)
        roi_corners = {
            'top_left': (x_min, y_max),
            'top_right': (x_max, y_max),
            'bottom_left': (x_min, y_min),
            'bottom_right': (x_max, y_min)
        }

    x, y, width, height = roi['x'], roi['y'], roi['width'], roi['height']
    
    if x < 0 or y < 0 or x + width > image.shape[1] or y + height > image.shape[0]:
        return jsonify({'error': 'ROI is out of bounds.'}), 400

    roi_image = image[y:y + height, x:x + width]
    
    # Apply Gaussian filter and Canny edge detection on the ROI
    gaussian_blurred = cv2.GaussianBlur(roi_image, (5, 5), 0)
    edges = cv2.Canny(gaussian_blurred, 100, 200)

    # Save processed images
    gaussian_image_path = os.path.join(UPLOAD_FOLDER, f'gaussian_{filename}')
    edges_image_path = os.path.join(UPLOAD_FOLDER, f'edges_{filename}')
    
    cv2.imwrite(gaussian_image_path, gaussian_blurred)
    cv2.imwrite(edges_image_path, edges)

    # Find contours from edges detected by Canny
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Draw contours on a blank image or the original ROI for visualization
    contour_image = np.zeros_like(roi_image)
    cv2.drawContours(contour_image, contours, -1, (0, 255, 0), 3)

    # Save the contour image
    contour_image_path = os.path.join(UPLOAD_FOLDER, f'contours_{filename}')
    cv2.imwrite(contour_image_path, contour_image)

    if roi_image.size > 0:
        roi_image_processed = preprocess_roi(roi_image)

        # Save the processed ROI image
        roi_image_filename = f'roi_{filename}'
        roi_image_path = os.path.join(UPLOAD_FOLDER, roi_image_filename)
        cv2.imwrite(roi_image_path, roi_image_processed)

        color_info, area_square_meters, avg_color_rgb, breadth, length, combined_path, common_color, surface_color_path, surface_color_base64 = analyze_colors(roi_image_processed)

        # Save stencil images and update paths
        dominant_colors = []
        for index, color in enumerate(color_info):
            
            dominant_colors.append({
                'Color' : '',
                'Hex': color['hex'],  # Use the correct key for background color
                'RGB': color['rgb'],
                'Stencil': '',
                'Path' : color['path'] # Use the saved path for stencil
            })

        # Prepare data for Excel
        excel_data = {
            'Average Color (RGB)': [avg_color_rgb],
            'Area (Pixels)': [width * height],
            'Area (Square Meters)': [area_square_meters],
            'Common Color': [common_color],
            'Latitude': [data.get('latitude')],
            'Longitude': [data.get('longitude')],
            'Breadth (Pixels)': [width],
            'Length (Pixels)': [height],
            'Breadth': [breadth],
            'Length': [length]
        }

        # Create a DataFrame for the main data
        main_df = pd.DataFrame(excel_data)

        # Create a DataFrame for dominant colors
        dominant_colors_df = pd.DataFrame(dominant_colors)

        # Save DataFrames to an Excel file
        excel_path = os.path.join(UPLOAD_FOLDER, f'patterns_data_{filename}.xlsx')
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            main_df.to_excel(writer, index=False, sheet_name='Main Data')
            dominant_colors_df.to_excel(writer, index=False, sheet_name='Dominant Colors')

            # Access the dominant colors sheet
            dominant_colors_sheet = writer.sheets['Dominant Colors']
            
            

            # Insert stencil images
            for index, color in enumerate(dominant_colors):
                print(color)
                fill_color = PatternFill(start_color=color['Hex'].replace('#', ''), end_color=color['Hex'].replace('#', ''), fill_type='solid')
                dominant_colors_sheet.cell(row=index + 2, column=1).fill = fill_color  # Start from row 2

                stencil_image_path = color['Path']  # Path to the stencil image
                if os.path.exists(stencil_image_path):  # Check if the file exists
                    try:
                        print('a')
                        img = OpenPyXLImage(stencil_image_path)
                        img.width = 100  # Adjust the image width if needed
                        img.height = 100  # Adjust the image height if needed
                        # Calculate the cell position (A5 is the starting point)
                        cell_row = 2 + index  # Starts from row 5 and increments
                        print(img)
                        dominant_colors_sheet.add_image(img, f'D{cell_row}')  # Insert image at A5, A6, etc.
                        print(f'Inserted image at D{cell_row}: {stencil_image_path}')
                        dominant_colors_sheet.row_dimensions[cell_row].height = 80  # Adjust as necessary

                    except Exception as e:
                        print(f'Error inserting image at A{cell_row}: {e}')
                else:
                    print(f'Image file not found: {stencil_image_path}')

                column_widths = {
            'A': 20,  # Width for Hex column
            'B': 20,  # Width for RGB column
            'C': 20,  # Width for Stencil column
            'D': 20, 
            'E' : 50  # Increased width for Image column to fit images
        }
        
                for column_letter, width in column_widths.items():
                     dominant_colors_sheet.column_dimensions[column_letter].width = width

        # Log the path of the generated Excel file
        print(f'Excel file created at: {excel_path}')

        return jsonify({
            'average_color': avg_color_rgb,
            'area_pixels': width * height,
            'area_square_meters': area_square_meters,
            'dominant_colors': color_info,
            'common_color': common_color,
            'roi_image_path': roi_image_path.replace('\\', '/'),
            'latitude': data.get('latitude'),
            'longitude': data.get('longitude'),
            'breadth_in_pixels': width,
            'length_in_pixels': height,
            'breadth': breadth,
            'length': length,
            'gaussian_image_path': gaussian_image_path.replace('\\', '/'),
            'edges_image_path': edges_image_path.replace('\\', '/'),
            'contour_image_path': contour_image_path.replace('\\', '/'),
            'excel_path': excel_path.replace('\\', '/') , # Return the path of the Excel file
            'roi_corners': roi_corners,
            'combined_path': combined_path.replace('\\', '/'),
            "surface_color_path" : surface_color_path.replace('\\', '/'),
            

        }), 200

def preprocess_roi(roi_image):
    """Preprocess the ROI: blur and sharpen."""
    roi_image_blurred = cv2.medianBlur(roi_image, 3)
    print("roi_image_blurred:" , roi_image_blurred)
    roi_image_resized = cv2.resize(roi_image_blurred, (roi_image.shape[1] * 2, roi_image.shape[0] * 2), interpolation=cv2.INTER_CUBIC)
    print("roi_image_resized:" , roi_image_resized)
    sharpening_kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
    print("sharpening_kernel:" , sharpening_kernel)
    return cv2.filter2D(roi_image_resized, -1, sharpening_kernel)


def compute_surface_color(roi_image):
    """Calculate the average color (surface color) of the entire ROI."""
    # Compute the average color of the ROI image
    mean_color = np.mean(roi_image, axis=(0, 1))  # Average over all pixels
    return mean_color.astype(int)


PIXELS_PER_METER = 100  # Example value; replace with the actual scaling factor if known.
def create_contour_mask(contour, image_shape):
    """Create a binary mask for a given contour with the same dimensions as the input image."""
    # Ensure shape is valid and extract only the (height, width)
    if isinstance(image_shape, tuple) and len(image_shape) >= 2:
        mask_shape = (int(image_shape[0]), int(image_shape[1]))  # Ensure these are integers
    else:
        raise ValueError("image_shape must be a tuple with at least two dimensions (height, width).")

    # Create the mask with the correct shape
    mask = np.zeros(mask_shape, dtype=np.uint8)
    cv2.drawContours(mask, [contour], -1, 255, thickness=cv2.FILLED)
    return mask

def process_stencil_for_color(contours, roi_image, dominant_color):
    """Generate a stencil image for a specific dominant color by filling contours that match this color."""
    stencil_image = np.zeros_like(roi_image)
    print(len(contours))
    for contour in contours:
        # Pass the shape of the image to ensure the mask dimensions match the input image
        contour_mask = create_contour_mask(contour, roi_image.shape)
        mean_color = cv2.mean(roi_image, mask=contour_mask)[:3]
        print(mean_color)
        
        if np.allclose(np.array(mean_color).astype(int), dominant_color, atol=30):
            cv2.drawContours(stencil_image, [contour], -1, dominant_color.tolist(), thickness=cv2.FILLED)
    return stencil_image

def rgb_to_hex(rgb):
    return "#{:02x}{:02x}{:02x}".format(*rgb)
def analyze_colors(roi_image, k=4):
    """Analyze colors in the ROI, generate and save stencils for each dominant color."""
    stencil_dir = "stencils"
    os.makedirs(f"{UPLOAD_FOLDER}/{stencil_dir}", exist_ok=True)
    print(f"Analyzing colors in {roi_image.shape}")
    # Resize for faster processing
    downsampled_image = cv2.resize(roi_image, (roi_image.shape[1] // 8, roi_image.shape[0] // 8))
    pixels = downsampled_image.reshape(-1, 3)
    print(f"Pixels shape: {pixels.shape}")
    # Cluster colors with KMeans
    kmeans = MiniBatchKMeans(n_clusters=k, init='k-means++', batch_size=1000)
    kmeans.fit(pixels)
    dominant_colors = kmeans.cluster_centers_.astype(int)
    print(f"Dominant colors: {dominant_colors}")

    # Prepare the image for contour detection
    gray_image = cv2.cvtColor(roi_image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(cv2.GaussianBlur(gray_image, (3, 3), 0), 127, 255, cv2.THRESH_BINARY)
    contours, _ = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    print(f"Contours: {contours}")
    
    # Calculate the surface color of the ROI (average color)
    surface_color = compute_surface_color(roi_image)
    print(f"Surface color: {surface_color}")

    # Create an image filled with the surface color
    surface_color_image = np.ones_like(roi_image) * 255  # White background
    surface_color_image[:, :] = surface_color 
    print(f"Surface color image: {surface_color_image.shape}")

    # Generate stencils for each dominant color
    color_info_list = []
    combined_image = surface_color_image.copy()  # Start with the surface color background
    
    for i, dominant_color in enumerate(dominant_colors):
        # Process stencil for each dominant color
        stencil_image = process_stencil_for_color(contours, roi_image, dominant_color)
        print(f"Stencil image shape: {stencil_image.shape}")
        print(i)
        
        # Overlay stencil on the combined image
        # We ensure that the surface color is kept for areas without stencils
        mask = stencil_image == dominant_color  # Where stencil matches dominant color
        combined_image[mask] = stencil_image[mask]  # Apply stencil color on the surface image

        # Save each stencil as an image file (for debugging)
        # stencil_path = os.path.join(f"{UPLOAD_FOLDER}/{stencil_dir}", f'stencil_{i}.png')
        # Image.fromarray(stencil_image).save(stencil_path)
        # Save each stencil as an image file
        stencil_path = os.path.join(f"{UPLOAD_FOLDER}/{stencil_dir}", f'stencil_{i}.png')
        Image.fromarray(stencil_image).save(stencil_path)
        print(f"Stencil saved to {stencil_path}")

        # Encode image for potential use in JSON
        buffered = io.BytesIO()
        Image.fromarray(stencil_image).save(buffered, format="PNG")
        stencil_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')

        color_info_list.append({
            'rgb': dominant_color.tolist(),
            'hex': rgb_to_hex(dominant_color),
            'stencil_image': stencil_base64,
            'path': stencil_path.replace('\\', '/')
        })

    # Save the surface color image and combined image
    surface_color_path = os.path.join(f"{UPLOAD_FOLDER}/{stencil_dir}", 'surface_color.png')
    Image.fromarray(surface_color_image).save(surface_color_path)

    combined_path = os.path.join(f"{UPLOAD_FOLDER}/{stencil_dir}", 'combined_stencil.png')
    Image.fromarray(combined_image).save(combined_path)

    # Encode surface color image for potential use in JSON
    surface_buffered = io.BytesIO()
    Image.fromarray(surface_color_image).save(surface_buffered, format="PNG")
    surface_color_base64 = base64.b64encode(surface_buffered.getvalue()).decode('utf-8')

    # Calculate the area in square meters based on pixel density
    area_square_meters = (roi_image.shape[0] * roi_image.shape[1]) / (PIXELS_PER_METER ** 2)

    # Calculate width and height in meters
    width = roi_image.shape[1] / PIXELS_PER_METER
    height = roi_image.shape[0] / PIXELS_PER_METER

    # Calculate average RGB intensity
    avg_color_rgb_intensity = np.mean(dominant_colors, axis=0).astype(int).tolist()

    # Determine the most common color
    labels = kmeans.labels_
    most_common_color_idx = np.argmax(np.bincount(labels))
    common_color = {
        'rgb': dominant_colors[most_common_color_idx].tolist(),
        'hex': rgb_to_hex(dominant_colors[most_common_color_idx])
    }

    return color_info_list, area_square_meters, avg_color_rgb_intensity, width, height, combined_path, common_color, surface_color_path, surface_color_base64

def find_optimal_n_clusters(inertias):
    # Calculate the first derivative (difference)
    first_derivative = np.diff(inertias)
    
    # Calculate the second derivative (change of the first derivative)
    second_derivative = np.diff(first_derivative)
    
    # The optimal cluster is at the index where the second derivative is maximized
    # We can find the maximum point in the second derivative
    optimal_index = np.argmax(second_derivative) + 1  # +1 because np.diff reduces the size by 1
    
    # Ensure optimal_index doesn't exceed the bounds of the original clusters
    optimal_n_clusters = np.clip(optimal_index, 1, len(inertias))
    
    return optimal_n_clusters


if __name__ == '__main__':
    app.run(debug=True,host = '0.0.0.0')